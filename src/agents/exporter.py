from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import Workbook, load_workbook


# =========================================================
# Helpers
# =========================================================

def _load_json(path: str | None) -> Dict[str, Any]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _safe_text(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _safe_float(x: Any, default: float = 0.0) -> float:
    try:
        return float(x)
    except Exception:
        return default


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    return wb


def _ensure_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def _auto_fit_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def _sort_qids(qid: str):
    a, b = qid.split(".")
    return (int(a), int(b))


def _append_dict_row(ws, row_dict: Dict[str, Any], headers: List[str]) -> None:
    ws.append([row_dict.get(h, "") for h in headers])


def _rewrite_sheet_with_headers(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def _sheet_is_empty(ws) -> bool:
    return ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None


# =========================================================
# Row builders
# =========================================================

def _build_student_summary_row(
    sheet_id: str,
    identifiers: Dict[str, Any],
    grades_payload: Dict[str, Any],
    qa_payload: Dict[str, Any],
    feedback_txt: str,
) -> Dict[str, Any]:
    student_id = _safe_text(identifiers.get("student_id", ""))
    student_name = _safe_text(identifiers.get("student_name", ""))

    grading = grades_payload.get("grading", {})
    totals = grading.get("totals", {})

    qa = qa_payload.get("qa", {})

    return {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "sheet_id": sheet_id,
        "student_id": student_id,
        "student_name": student_name,
        "q1_total": _safe_float(totals.get("q1", 0)),
        "q2_total": _safe_float(totals.get("q2", 0)),
        "q3_total": _safe_float(totals.get("q3", 0)),
        "q4_total": _safe_float(totals.get("q4", 0)),
        "overall_total": _safe_float(totals.get("overall", 0)),
        "qa_status": _safe_text(qa.get("status", "")),
        "qa_flag_count": _safe_float(qa.get("flag_count", 0)),
        "qa_questions_to_review": ", ".join(qa.get("questions_to_review", []) or []),
        "feedback_available": bool(feedback_txt.strip()),
    }


def _build_question_detail_rows(
    sheet_id: str,
    identifiers: Dict[str, Any],
    grades_payload: Dict[str, Any],
) -> List[Dict[str, Any]]:
    student_id = _safe_text(identifiers.get("student_id", ""))
    student_name = _safe_text(identifiers.get("student_name", ""))

    grading = grades_payload.get("grading", {})
    questions = grading.get("questions", {})

    rows: List[Dict[str, Any]] = []

    for qid in sorted(questions.keys(), key=_sort_qids):
        qobj = questions.get(qid, {})
        rows.append({
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "sheet_id": sheet_id,
            "student_id": student_id,
            "student_name": student_name,
            "question_id": qid,
            "answer": _safe_text(qobj.get("answer", "")),
            "score": _safe_float(qobj.get("score", 0)),
            "max_score": _safe_float(qobj.get("max_score", 0)),
            "reason": _safe_text(qobj.get("reason", "")),
        })

    return rows


# =========================================================
# Main API
# =========================================================

def export_student_xlsx(
    sheet_id: str,
    grades_json_path: str,
    identifiers_json_path: str | None,
    out_xlsx_path: str,
    qa_json_path: str | None = None,
    feedback_txt_path: str | None = None,
) -> str:
    grades_payload = _load_json(grades_json_path)
    identifiers = _load_json(identifiers_json_path)
    qa_payload = _load_json(qa_json_path)

    feedback_txt = ""
    if feedback_txt_path:
        p = Path(feedback_txt_path)
        if p.exists():
            feedback_txt = p.read_text(encoding="utf-8").strip()

    summary_row = _build_student_summary_row(
        sheet_id=sheet_id,
        identifiers=identifiers,
        grades_payload=grades_payload,
        qa_payload=qa_payload,
        feedback_txt=feedback_txt,
    )

    question_rows = _build_question_detail_rows(
        sheet_id=sheet_id,
        identifiers=identifiers,
        grades_payload=grades_payload,
    )

    out_path = Path(out_xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = _ensure_workbook(out_path)

    # =====================================================
    # Sheet 1: students
    # =====================================================
    if "students" in wb.sheetnames:
        ws_students = wb["students"]
    else:
        ws_students = wb.active
        ws_students.title = "students"

    student_headers = [
        "exported_at",
        "sheet_id",
        "student_id",
        "student_name",
        "q1_total",
        "q2_total",
        "q3_total",
        "q4_total",
        "overall_total",
        "qa_status",
        "qa_flag_count",
        "qa_questions_to_review",
        "feedback_available",
    ]

    if _sheet_is_empty(ws_students):
        ws_students.append(student_headers)
        _append_dict_row(ws_students, summary_row, student_headers)
    else:
        existing_headers = [cell.value for cell in ws_students[1]]
        if existing_headers != student_headers:
            existing_rows = []
            for row in ws_students.iter_rows(min_row=2, values_only=True):
                existing_rows.append(dict(zip(existing_headers, row)))
            existing_rows.append(summary_row)
            _rewrite_sheet_with_headers(ws_students, student_headers, existing_rows)
        else:
            _append_dict_row(ws_students, summary_row, student_headers)

    _auto_fit_width(ws_students)

    # =====================================================
    # Sheet 2: question_details
    # =====================================================
    ws_details = _ensure_sheet(wb, "question_details")

    detail_headers = [
        "exported_at",
        "sheet_id",
        "student_id",
        "student_name",
        "question_id",
        "answer",
        "score",
        "max_score",
        "reason",
    ]

    if _sheet_is_empty(ws_details):
        ws_details.append(detail_headers)
        for row_dict in question_rows:
            _append_dict_row(ws_details, row_dict, detail_headers)
    else:
        existing_headers = [cell.value for cell in ws_details[1]]
        if existing_headers != detail_headers:
            existing_rows = []
            for row in ws_details.iter_rows(min_row=2, values_only=True):
                existing_rows.append(dict(zip(existing_headers, row)))
            existing_rows.extend(question_rows)
            _rewrite_sheet_with_headers(ws_details, detail_headers, existing_rows)
        else:
            for row_dict in question_rows:
                _append_dict_row(ws_details, row_dict, detail_headers)

    _auto_fit_width(ws_details)

    wb.save(out_path)
    return str(out_path)